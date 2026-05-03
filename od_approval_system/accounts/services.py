import csv
from dataclasses import dataclass, field
from io import TextIOWrapper
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from .models import Department, FacultyProfile, School, StudentProfile, UserRole

User = get_user_model()

ALLOWED_STUDENT_IMPORT_EXTENSIONS = {'.csv', '.xlsx'}
BASE_REQUIRED_HEADERS = {'email', 'roll_number', 'password'}
OPTIONAL_HEADERS = {
    'first_name', 'last_name', 'mobile', 'school_code', 'department_code',
    'year', 'section', 'semester', 'mentor_email'
}


@dataclass
class StudentImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)

    @property
    def total_success(self):
        return self.created + self.updated


def read_student_upload_rows(uploaded_file):
    """Read CSV/XLSX uploads into normalized dictionaries."""
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix not in ALLOWED_STUDENT_IMPORT_EXTENSIONS:
        raise ValidationError('Only CSV and Excel .xlsx files are allowed for student uploads.')

    if suffix == '.xlsx':
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or '').strip().lower() for h in rows[0]]
        return [
            {headers[i]: (row[i] if i < len(row) else '') for i in range(len(headers))}
            for row in rows[1:]
            if any(cell not in (None, '') for cell in row)
        ]

    uploaded_file.seek(0)
    wrapper = TextIOWrapper(uploaded_file.file, encoding='utf-8-sig')
    return [dict(row) for row in csv.DictReader(wrapper)]


def _clean_value(row, key, default=''):
    value = row.get(key, default)
    if value is None:
        return default
    return str(value).strip()


def _int_value(row, key, default):
    raw = _clean_value(row, key, default)
    try:
        return int(raw)
    except (TypeError, ValueError):
        return default


def _resolve_school_department(row, uploaded_by, scope):
    if scope == 'faculty':
        profile = uploaded_by.faculty_profile
        return profile.school, profile.department

    if scope == 'dean':
        school = uploaded_by.dean_profile.school
        department_code = _clean_value(row, 'department_code')
        if not department_code:
            raise ValidationError('department_code is required for dean uploads.')
        department = Department.objects.get(code__iexact=department_code, school=school, is_active=True)
        return school, department

    school_code = _clean_value(row, 'school_code')
    department_code = _clean_value(row, 'department_code')
    if not school_code or not department_code:
        raise ValidationError('school_code and department_code are required for admin uploads.')
    school = School.objects.get(code__iexact=school_code, is_active=True)
    department = Department.objects.get(code__iexact=department_code, school=school, is_active=True)
    return school, department


def _resolve_mentor(row, uploaded_by, scope, school, department):
    if scope == 'faculty':
        return uploaded_by

    mentor_email = _clean_value(row, 'mentor_email').lower()
    if mentor_email:
        try:
            return User.objects.get(
                email__iexact=mentor_email,
                role=UserRole.FACULTY,
                is_active=True,
                faculty_profile__school=school,
                faculty_profile__department=department,
                faculty_profile__is_active=True,
            )
        except User.DoesNotExist:
            raise ValidationError(f'Mentor faculty not found or inactive: {mentor_email}')

    if scope == 'dean':
        return None
    return None


def import_student_credentials(uploaded_file, uploaded_by, scope='admin'):
    """
    Create/update student users from staff-uploaded credentials.

    The uploaded password is never stored as plain text. It is immediately hashed
    through Django's password hasher and attached to the student user account.
    """
    rows = read_student_upload_rows(uploaded_file)
    result = StudentImportResult()
    if not rows:
        result.errors.append('The uploaded file is empty.')
        return result

    normalized_rows = []
    for row in rows:
        normalized_rows.append({str(k or '').strip().lower(): v for k, v in row.items()})

    missing = BASE_REQUIRED_HEADERS - set(normalized_rows[0].keys())
    if missing:
        result.errors.append('Missing required column(s): ' + ', '.join(sorted(missing)))
        return result

    for index, row in enumerate(normalized_rows, start=2):
        email = _clean_value(row, 'email').lower()
        roll_number = _clean_value(row, 'roll_number').upper()
        password = _clean_value(row, 'password')
        if not email or not roll_number or not password:
            result.skipped += 1
            result.errors.append(f'Row {index}: email, roll_number and password are mandatory.')
            continue
        if len(password) < 8:
            result.skipped += 1
            result.errors.append(f'Row {index}: password must be at least 8 characters.')
            continue

        try:
            school, department = _resolve_school_department(row, uploaded_by, scope)
            mentor = _resolve_mentor(row, uploaded_by, scope, school, department)
        except Exception as exc:
            result.skipped += 1
            result.errors.append(f'Row {index}: {exc}')
            continue

        existing_roll = StudentProfile.objects.filter(roll_number__iexact=roll_number).select_related('user').first()
        existing_user = User.objects.filter(email__iexact=email).first()
        if existing_roll and (not existing_user or existing_roll.user_id != existing_user.id):
            result.skipped += 1
            result.errors.append(f'Row {index}: roll number {roll_number} belongs to another email.')
            continue
        if existing_user and existing_user.role != UserRole.STUDENT:
            result.skipped += 1
            result.errors.append(f'Row {index}: {email} already belongs to a non-student user.')
            continue

        with transaction.atomic():
            user_created = existing_user is None
            user = existing_user or User(email=email, username=email, role=UserRole.STUDENT)
            user.email = email
            user.username = email
            user.role = UserRole.STUDENT
            user.first_name = _clean_value(row, 'first_name', user.first_name or 'Student') or 'Student'
            user.last_name = _clean_value(row, 'last_name', user.last_name or '')
            user.mobile = _clean_value(row, 'mobile', user.mobile or '')
            user.is_active = True
            user.must_change_password = False
            user.set_password(password)
            user.save()

            defaults = {
                'roll_number': roll_number,
                'school': school,
                'department': department,
                'year': _int_value(row, 'year', 1),
                'section': _clean_value(row, 'section', 'A') or 'A',
                'semester': _int_value(row, 'semester', 1),
                'mentor': mentor,
                'staff_uploaded_login': True,
                'uploaded_by': uploaded_by,
                'uploaded_at': timezone.now(),
                'is_active': True,
            }
            StudentProfile.objects.update_or_create(user=user, defaults=defaults)

        if user_created:
            result.created += 1
        else:
            result.updated += 1

    return result


@dataclass
class StudentImportPreviewRow:
    row_number: int
    email: str
    roll_number: str
    name: str
    school_code: str
    department_code: str
    status: str
    message: str


def preview_student_credentials(uploaded_file, uploaded_by, scope='admin'):
    """Validate student upload without writing to the database."""
    rows = read_student_upload_rows(uploaded_file)
    preview = []
    if not rows:
        return [StudentImportPreviewRow(1, '', '', '', '', '', 'error', 'The uploaded file is empty.')]
    normalized_rows = [{str(k or '').strip().lower(): v for k, v in row.items()} for row in rows]
    missing = BASE_REQUIRED_HEADERS - set(normalized_rows[0].keys())
    if missing:
        return [StudentImportPreviewRow(1, '', '', '', '', '', 'error', 'Missing required column(s): ' + ', '.join(sorted(missing)))]

    seen_emails = set()
    seen_rolls = set()
    for index, row in enumerate(normalized_rows, start=2):
        email = _clean_value(row, 'email').lower()
        roll_number = _clean_value(row, 'roll_number').upper()
        password = _clean_value(row, 'password')
        school_code = _clean_value(row, 'school_code')
        department_code = _clean_value(row, 'department_code')
        name = (_clean_value(row, 'first_name') + ' ' + _clean_value(row, 'last_name')).strip()
        status, message = 'ok', 'Ready to import'
        if not email or not roll_number or not password:
            status, message = 'error', 'email, roll_number, and password are mandatory'
        elif email in seen_emails:
            status, message = 'error', f'Duplicate email in file: {email}'
        elif roll_number in seen_rolls:
            status, message = 'error', f'Duplicate roll number in file: {roll_number}'
        elif len(password) < 8:
            status, message = 'error', 'Password must be at least 8 characters'
        else:
            try:
                school, department = _resolve_school_department(row, uploaded_by, scope)
                _resolve_mentor(row, uploaded_by, scope, school, department)
                existing_roll = StudentProfile.objects.filter(roll_number__iexact=roll_number).select_related('user').first()
                existing_user = User.objects.filter(email__iexact=email).first()
                if existing_roll and (not existing_user or existing_roll.user_id != existing_user.id):
                    status, message = 'error', f'Roll number {roll_number} belongs to another email'
                elif existing_user and existing_user.role != UserRole.STUDENT:
                    status, message = 'error', f'{email} belongs to a non-student user'
                elif existing_user:
                    status, message = 'update', 'Existing student will be updated and password reset'
                else:
                    status, message = 'create', 'New student account will be created'
            except Exception as exc:
                status, message = 'error', str(exc)
        seen_emails.add(email)
        seen_rolls.add(roll_number)
        preview.append(StudentImportPreviewRow(index, email, roll_number, name, school_code, department_code, status, message))
    return preview
